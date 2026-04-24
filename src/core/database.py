import sqlite3
import pandas as pd
import openpyxl
from datetime import datetime
from src.core.config import EXCEL_DB_PATH, SQLITE_DB_PATH
from src.core.safety import SafetyManager
import os
import logging

logger = logging.getLogger(__name__)

class DatabaseEngine:
    def __init__(self):
        self.sqlite_path = SQLITE_DB_PATH
        self.excel_path = EXCEL_DB_PATH
        self._init_sqlite()

    def _init_sqlite(self):
        """Initialize SQLite schema if it doesn't exist."""
        conn = sqlite3.connect(self.sqlite_path)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                product_id TEXT PRIMARY KEY,
                sku_code TEXT,
                name TEXT,
                category TEXT,
                unit TEXT,
                status TEXT,
                sell_price REAL,
                cost_price REAL,
                reorder_qty INTEGER
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS purchases (
                purchase_id TEXT PRIMARY KEY,
                date TEXT,
                product_id TEXT,
                batch_id TEXT,
                supplier TEXT DEFAULT '',
                qty INTEGER,
                cost_per_unit REAL,
                total_cost REAL,
                FOREIGN KEY (product_id) REFERENCES products (product_id)
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                sales_id TEXT PRIMARY KEY,
                date TEXT,
                product_id TEXT,
                customer TEXT DEFAULT 'Walk-in',
                qty INTEGER,
                sell_price REAL,
                discount REAL DEFAULT 0,
                revenue REAL,
                cogs REAL,
                profit REAL,
                FOREIGN KEY (product_id) REFERENCES products (product_id)
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                user TEXT,
                action TEXT,
                details TEXT
            )
        ''')

        # Indexes for performance on 50,000+ records
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_sales_product ON sales(product_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_purchases_product ON purchases(product_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_purchases_date ON purchases(date)')

        conn.commit()
        conn.close()

    def sync_from_excel(self):
        """Sync Excel data into SQLite cache."""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel database not found at {self.excel_path}")

        conn = sqlite3.connect(self.sqlite_path)
        
        try:
            # Sync Products
            df_products = pd.read_excel(self.excel_path, sheet_name='Product_Master', skiprows=3)
            df_products = df_products[['Product ID', 'SKU Code', 'Product Name', 'Category', 'Unit', 'Status', 'Sell Price ৳', 'Cost ৳ / Unit', 'Reorder Qty']]
            df_products.columns = ['product_id', 'sku_code', 'name', 'category', 'unit', 'status', 'sell_price', 'cost_price', 'reorder_qty']
            df_products.dropna(subset=['product_id'], inplace=True)
            df_products.to_sql('products', conn, if_exists='replace', index=False)

            # Sync Purchases (Purchase ID is a formula → filter by Date instead)
            df_purchases = pd.read_excel(self.excel_path, sheet_name='Purchase_Log', skiprows=3)
            df_purchases = df_purchases[['Purchase ID', 'Date', 'Product ID', 'Batch ID', 'Qty Purchased', 'Cost / Unit ৳', 'Total Cost ৳']]
            df_purchases.columns = ['purchase_id', 'date', 'product_id', 'batch_id', 'qty', 'cost_per_unit', 'total_cost']
            df_purchases.dropna(subset=['date'], inplace=True)
            # Cast formula columns to string so we can assign generated IDs
            df_purchases['purchase_id'] = df_purchases['purchase_id'].astype(object)
            df_purchases['batch_id'] = df_purchases['batch_id'].astype(object)
            for i, row in df_purchases.iterrows():
                if pd.isna(row['purchase_id']) or str(row['purchase_id']).strip() == '':
                    df_purchases.at[i, 'purchase_id'] = f"PUR-{i+1:04d}"
                if pd.isna(row['batch_id']) or str(row['batch_id']).strip() == '':
                    df_purchases.at[i, 'batch_id'] = f"{row['product_id']}-BATCH-{i+1}"
                if pd.isna(row['total_cost']):
                    df_purchases.at[i, 'total_cost'] = (row['qty'] or 0) * (row['cost_per_unit'] or 0)
            df_purchases['date'] = df_purchases['date'].astype(str)
            df_purchases['supplier'] = ''
            df_purchases.to_sql('purchases', conn, if_exists='replace', index=False)

            # Sync Sales (Sales ID is a formula → filter by Date instead)
            df_sales = pd.read_excel(self.excel_path, sheet_name='Sales_Log', skiprows=3)
            df_sales = df_sales[['Sales ID', 'Date', 'Product ID', 'Qty Sold', 'Sell Price ৳', 'Revenue ৳', 'COGS ৳', 'Profit ৳']]
            df_sales.columns = ['sales_id', 'date', 'product_id', 'qty', 'sell_price', 'revenue', 'cogs', 'profit']
            df_sales.dropna(subset=['date'], inplace=True)
            df_sales['sales_id'] = df_sales['sales_id'].astype(object)
            for i, row in df_sales.iterrows():
                if pd.isna(row['sales_id']) or str(row['sales_id']).strip() == '':
                    df_sales.at[i, 'sales_id'] = f"SL-{i+1:05d}"
                for col in ['revenue', 'cogs', 'profit']:
                    if pd.isna(df_sales.at[i, col]):
                        df_sales.at[i, col] = 0
            df_sales['date'] = df_sales['date'].astype(str)
            df_sales['customer'] = 'Walk-in'
            df_sales['discount'] = 0
            df_sales.to_sql('sales', conn, if_exists='replace', index=False)

            conn.commit()
            logger.info("Excel -> SQLite sync completed successfully.")
            print("Sync complete.")
        except Exception as e:
            logger.error(f"Sync failed: {e}")
            raise
        finally:
            conn.close()

    def execute_query(self, query, params=()):
        """Execute a SELECT query and return a pandas DataFrame."""
        conn = sqlite3.connect(self.sqlite_path)
        try:
            df = pd.read_sql_query(query, conn, params=params)
            return df
        finally:
            conn.close()

    def execute_write(self, query, params=()):
        """Execute an INSERT/UPDATE/DELETE query."""
        conn = sqlite3.connect(self.sqlite_path)
        cursor = conn.cursor()
        try:
            cursor.execute(query, params)
            conn.commit()
        finally:
            conn.close()

    def get_next_sale_id(self):
        """Generate a unique sequential sales ID."""
        df = self.execute_query("SELECT COUNT(*) as cnt FROM sales")
        count = int(df.iloc[0]['cnt']) + 1
        return f"SL-{count:05d}"

    def get_next_purchase_id(self):
        """Generate a unique sequential purchase ID."""
        df = self.execute_query("SELECT COUNT(*) as cnt FROM purchases")
        count = int(df.iloc[0]['cnt']) + 1
        return f"PUR-{count:04d}"

    def write_sale(self, sale_data):
        """Write a sale to both SQLite and Excel."""
        SafetyManager.create_backup()

        query = """INSERT INTO sales 
            (sales_id, date, product_id, customer, qty, sell_price, discount, revenue, cogs, profit) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
        params = (
            sale_data['sales_id'], sale_data['date'], sale_data['product_id'],
            sale_data.get('customer', 'Walk-in'), sale_data['qty'], sale_data['sell_price'],
            sale_data.get('discount', 0), sale_data['revenue'],
            sale_data['cogs'], sale_data['profit']
        )
        self.execute_write(query, params)

        # Log audit
        self.log_audit("SYSTEM", "SALE", f"Sale {sale_data['sales_id']} recorded for {sale_data.get('customer', 'Walk-in')}")

        # Write to Excel if auto-sync is enabled
        from src.core.config import AUTO_SYNC_EXCEL
        if not AUTO_SYNC_EXCEL:
            return

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Sales_Log']
            
            row = 5
            while ws.cell(row=row, column=2).value is not None:
                row += 1
                
            ws.cell(row=row, column=1).value = f'=IF(B{row}="","","SL-"&TEXT({row-4},"00000"))'
            ws.cell(row=row, column=2).value = datetime.strptime(sale_data['date'], '%Y-%m-%d')
            ws.cell(row=row, column=3).value = sale_data['product_id']
            ws.cell(row=row, column=4).value = f'=IFERROR(INDEX(pmName,MATCH(C{row},pmID,0)),"")'
            ws.cell(row=row, column=5).value = sale_data['qty']
            ws.cell(row=row, column=6).value = sale_data['sell_price']
            ws.cell(row=row, column=7).value = f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})'
            
            wb.save(self.excel_path)
            logger.info(f"Sale {sale_data['sales_id']} written to Excel and SQLite.")
        except Exception as e:
            logger.error(f"Excel write failed for sale {sale_data['sales_id']}: {e}")

    def write_purchase(self, purchase_data):
        """Write a purchase to both SQLite and Excel."""
        SafetyManager.create_backup()

        query = """INSERT INTO purchases 
            (purchase_id, date, product_id, batch_id, supplier, qty, cost_per_unit, total_cost) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)"""
        params = (
            purchase_data['purchase_id'], purchase_data['date'], purchase_data['product_id'],
            purchase_data['batch_id'], purchase_data.get('supplier', ''),
            purchase_data['qty'], purchase_data['cost_per_unit'],
            purchase_data['qty'] * purchase_data['cost_per_unit']
        )
        self.execute_write(query, params)

        # Log audit
        self.log_audit("SYSTEM", "PURCHASE", f"Purchase {purchase_data['purchase_id']} recorded for product {purchase_data['product_id']}")

        # Write to Excel if auto-sync is enabled
        from src.core.config import AUTO_SYNC_EXCEL
        if not AUTO_SYNC_EXCEL:
            return

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Purchase_Log']
            
            row = 5
            while ws.cell(row=row, column=2).value is not None:
                row += 1
                
            ws.cell(row=row, column=1).value = f'=IF(B{row}="","","PUR-"&TEXT({row-4},"0000"))'
            ws.cell(row=row, column=2).value = datetime.strptime(purchase_data['date'], '%Y-%m-%d')
            ws.cell(row=row, column=3).value = purchase_data['product_id']
            ws.cell(row=row, column=5).value = purchase_data['batch_id']
            ws.cell(row=row, column=6).value = purchase_data['qty']
            ws.cell(row=row, column=7).value = purchase_data['cost_per_unit']
            
            wb.save(self.excel_path)
            logger.info(f"Purchase {purchase_data['purchase_id']} written to Excel and SQLite.")
        except Exception as e:
            logger.error(f"Excel write failed for purchase {purchase_data['purchase_id']}: {e}")

    def log_audit(self, user, action, details):
        """Write an audit log entry."""
        self.execute_write(
            "INSERT INTO audit_logs (timestamp, user, action, details) VALUES (?, ?, ?, ?)",
            (datetime.now().isoformat(), user, action, details)
        )

    def sync_to_excel(self):
        """Manually sync all SQLite data back to the Master Excel file."""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Sync Sales
            ws_sales = wb['Sales_Log']
            df_sales = self.execute_query("SELECT * FROM sales")
            # Clear old rows (starting from row 5)
            for row in range(5, ws_sales.max_row + 1):
                for col in range(1, 10):
                    ws_sales.cell(row=row, column=col).value = None
            
            for i, row in df_sales.iterrows():
                r = i + 5
                ws_sales.cell(row=r, column=1).value = row['sales_id']
                ws_sales.cell(row=r, column=2).value = row['date']
                ws_sales.cell(row=r, column=3).value = row['product_id']
                ws_sales.cell(row=r, column=5).value = row['qty']
                ws_sales.cell(row=r, column=6).value = row['sell_price']
                ws_sales.cell(row=r, column=7).value = row['revenue']
            
            wb.save(self.excel_path)
            return True, "Sync to Excel completed successfully!"
        except Exception as e:
            return False, str(e)
