import os
import logging
import sqlite3
import pandas as pd
import openpyxl
from datetime import datetime
from src.core.config import EXCEL_DB_PATH, SQLITE_DB_PATH, AUTO_SYNC_EXCEL

logger = logging.getLogger(__name__)

class SyncManager:
    """Manages synchronization between Excel Master Database and SQLite Cache."""
    
    def __init__(self, database_engine):
        self.db = database_engine
        self.excel_path = EXCEL_DB_PATH
        self.sqlite_path = SQLITE_DB_PATH

    def sync_all_from_excel(self):
        """Full rebuild of SQLite cache from Excel master."""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel database not found at {self.excel_path}")

        conn = sqlite3.connect(self.sqlite_path)
        
        try:
            # Sync Products
            df_products = pd.read_excel(self.excel_path, sheet_name='Product_Master', skiprows=3)
            df_products = df_products[['Product ID', 'SKU Code', 'Product Name', 'Category', 'Unit', 'Status', 'Sell Price ৳', 'Cost ৳ / Unit', 'Reorder Qty']]
            df_products.columns = ['product_id', 'sku_code', 'name', 'category', 'unit', 'status', 'sell_price', 'cost_price', 'reorder_qty']
            df_products.dropna(subset=['product_id'], inplace=True)
            df_products.to_sql('products', conn, if_exists='replace', index=False)

            # Sync Purchases
            df_purchases = pd.read_excel(self.excel_path, sheet_name='Purchase_Log', skiprows=3)
            df_purchases = df_purchases[['Purchase ID', 'Date', 'Product ID', 'Batch ID', 'Qty Purchased', 'Cost / Unit ৳', 'Total Cost ৳']]
            df_purchases.columns = ['purchase_id', 'date', 'product_id', 'batch_id', 'qty', 'cost_per_unit', 'total_cost']
            df_purchases.dropna(subset=['date'], inplace=True)
            
            # ID generation for Excel formulas
            for i, row in df_purchases.iterrows():
                if pd.isna(row['purchase_id']) or str(row['purchase_id']).strip() == '':
                    df_purchases.at[i, 'purchase_id'] = f"PUR-{i+1:04d}"
                if pd.isna(row['batch_id']) or str(row['batch_id']).strip() == '':
                    df_purchases.at[i, 'batch_id'] = f"{row['product_id']}-BATCH-{i+1}"
            
            df_purchases['date'] = df_purchases['date'].astype(str)
            df_purchases['supplier'] = 'Standard'
            df_purchases.to_sql('purchases', conn, if_exists='replace', index=False)

            # Sync Sales
            df_sales = pd.read_excel(self.excel_path, sheet_name='Sales_Log', skiprows=3)
            df_sales = df_sales[['Sales ID', 'Date', 'Product ID', 'Qty Sold', 'Sell Price ৳', 'Revenue ৳', 'COGS ৳', 'Profit ৳']]
            df_sales.columns = ['sales_id', 'date', 'product_id', 'qty', 'sell_price', 'revenue', 'cogs', 'profit']
            df_sales.dropna(subset=['date'], inplace=True)
            
            for i, row in df_sales.iterrows():
                if pd.isna(row['sales_id']) or str(row['sales_id']).strip() == '':
                    df_sales.at[i, 'sales_id'] = f"SL-{i+1:05d}"
            
            df_sales['date'] = df_sales['date'].astype(str)
            df_sales['customer'] = 'Walk-in'
            df_sales['discount'] = 0
            df_sales.to_sql('sales', conn, if_exists='replace', index=False)

            conn.commit()
            logger.info("Excel -> SQLite sync completed successfully.")
            return True, "Sync complete."
        except Exception as e:
            logger.error(f"Sync failed: {e}")
            return False, str(e)
        finally:
            conn.close()

    def sync_sale_to_excel(self, sale_data: dict):
        """Append a single sale record to the Excel database."""
        if not AUTO_SYNC_EXCEL:
            return True
            
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Sales_Log']
            
            row = 5
            while ws.cell(row=row, column=2).value is not None:
                row += 1
                
            ws.cell(row=row, column=1).value = f'=IF(B{row}="","","SL-"&TEXT({row-4},"00000"))'
            ws.cell(row=row, column=2).value = datetime.strptime(sale_data['date'], '%Y-%m-%d')
            ws.cell(row=row, column=3).value = sale_data['product_id']
            ws.cell(row=row, column=4).value = f'=IFERROR(INDEX(pmName,MATCH(C{row},pmID,0)),"")'
            ws.cell(row=row, column=5).value = sale_data['qty']
            ws.cell(row=row, column=6).value = sale_data['sell_price']
            ws.cell(row=row, column=7).value = f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})'
            
            wb.save(self.excel_path)
            logger.info(f"Sale {sale_data.get('sale_id', 'N/A')} synced to Excel.")
            return True
        except Exception as e:
            logger.error(f"Excel sync failed for sale: {e}")
            return False

    def sync_purchase_to_excel(self, purchase_data: dict):
        """Append a single purchase record to the Excel database."""
        if not AUTO_SYNC_EXCEL:
            return True
            
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Purchase_Log']
            
            row = 5
            while ws.cell(row=row, column=2).value is not None:
                row += 1
                
            ws.cell(row=row, column=1).value = f'=IF(B{row}="","","PUR-"&TEXT({row-4},"0000"))'
            ws.cell(row=row, column=2).value = datetime.strptime(purchase_data['date'], '%Y-%m-%d')
            ws.cell(row=row, column=3).value = purchase_data['product_id']
            ws.cell(row=row, column=5).value = purchase_data['batch_id']
            ws.cell(row=row, column=6).value = purchase_data['qty']
            ws.cell(row=row, column=7).value = purchase_data['cost_per_unit']
            
            wb.save(self.excel_path)
            logger.info(f"Purchase {purchase_data.get('purchase_id', 'N/A')} synced to Excel.")
            return True
        except Exception as e:
            logger.error(f"Excel sync failed for purchase: {e}")
            return False

    def full_sync_to_excel(self):
        """Manual trigger to sync all SQLite data back to Excel."""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Sync Sales
            ws_sales = wb['Sales_Log']
            df_sales = self.db.execute_query("SELECT * FROM sales")
            # Clear old rows
            for r in range(5, ws_sales.max_row + 1):
                for c in range(1, 10):
                    ws_sales.cell(row=r, column=c).value = None
            
            for i, row in df_sales.iterrows():
                r = i + 5
                ws_sales.cell(row=r, column=1).value = row['sale_id']
                ws_sales.cell(row=r, column=2).value = row['date']
                ws_sales.cell(row=r, column=3).value = row['product_id']
                ws_sales.cell(row=r, column=5).value = row['qty']
                ws_sales.cell(row=r, column=6).value = row['sell_price']
                ws_sales.cell(row=r, column=7).value = row['revenue']
            
            wb.save(self.excel_path)
            return True, "Sync to Excel completed successfully!"
        except Exception as e:
            logger.error(f"Full sync to Excel failed: {e}")
            return False, str(e)
